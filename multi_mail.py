import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('Email_list.xlsx')
sheet = wb['Sheet1']

# lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=4).value

multiMail = {}
marker = "AUNIQUEMARKER"

sender = "Khuong Minh <khuongmi@hcsacc.education>"
# receiver = "Lu Khuong <khuonglu1999@gmail.com"

for r in range(2,3):
    name = sheet.cell(row=r, column=1).value
    email = sheet.cell(row=r, column=2).value
    multiMail[name] = email

smtpObj = smtplib.SMTP('smtp.outlook.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("khuongmi@hcsacc.education", sys.argv[1])

for name,email in multiMail.items():
    # header = """From: From Person <khuongmi@hcsacc.education>
    # To: To Person <%s>
    # Subject: Friendly reminder of today.
    # Content-Type: multipart/mixed; boundary=%s
    # --%s
    # """ % (email, marker, marker)

    body = """Subject: Friendly reminder of today.\t\t\tTo: khuong lu <khuonglu1999@gmail.com>\n
    Dear %s,
    Records show that you have not complete taskes dues for today.
    Please do the taskes as soon as possible.
    Thank you,
    Best regards,
    Khuong Minh Lu \n\n
    Vietname-German-University
    Tel: 0906507761
    Student-ID: 13156 CS2017 """  %( name)

    # mess = header + body

    print('Sending email to %s...' % email)

    sendmailStatus = smtpObj.sendmail(sender, email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
